# version 0.2.
import logging
from datetime import datetime

from openpyxl import load_workbook
from pyad import pyad_setdefaults, aduser, adcontainer
from pyad.adcontainer import ADContainer
from transliterate import translit
import ttkbootstrap as ttk
from ttkbootstrap.dialogs import Messagebox

from module.forms import *

LAST_VALUES = 'data.pkl'

# Setup gettext
LOCALE_DIR = 'locale'  # Directory containing locale files
LANGUAGE_CODE = 'en_US'  # Default language code


def is_user_exists(uname):
    """
    Define whether a user exists in the Active Directory.
    Args:
        uname(str): Username to check.

    Returns:
        bool: True if the user exists in the Active Directory.

    """
    try:
        aduser.ADUser.from_cn(uname)
        return True
    except Exception as e:
        return False


def create_ad_user(ou, upn_suffix, **kwargs):
    """
    Create an Active Directory user.
    Args:
        ou(ADContainer): Destination where the user should be created.
        upn_suffix(str): UPN-suffix for the domain where the user should be created.
        **kwargs: Additional keyword arguments for new user. Should at least contain 'cname', 'full_name', 'passwd', 'email'.

    Returns: None

    """
    try:
        if is_user_exists(uname=kwargs['cname']):
            if FIRST_RUN:
                logging.error(
                    f'Unable to create User {kwargs["cname"]} as same cname already exists! Change cname for user!')
                raise ValueError(
                    'Duplicate username detected. Change cname for user!')
            print(f'User "{kwargs['cname']}" already exists')
            logging.info(f'User "{kwargs['cname']}" already exists')
            return
        else:
            new_user = aduser.ADUser.create(name=kwargs['cname'], container_object=ou, upn_suffix=upn_suffix,
                                            password=kwargs['passwd'])
            new_user.update_attributes({
                'displayName': kwargs['full_name'],
                'mail': kwargs['email'],
                'userPrincipalName': f"{kwargs['cname']}@{upn_suffix}",
                'sn': kwargs['surname'],
                'givenName': str(kwargs['full_name'].split()[1]),
                'telephoneNumber': str(kwargs.get('phone')),
            })
            new_user.force_pwd_change_on_login()
            logging.info(f'User "{new_user.name}" created successfully')
    except Exception as e:
        print('Following error occurred: ', e)
        print(f'for user {kwargs['cname']}')
        logging.error(f'User {kwargs['cname']} was not created: {e}')
        raise e


def get_or_create_ou(ou_name, p_dn) -> ADContainer:
    """
    Return OU from the Active Directory if exists. If not exists, create it.
    Args:
        ou_name(str): Name of OU.
        p_dn(str): Parent DN for the OU to be created if not exists.

    Returns:
        ADContainer: Created or fetched OU.

    """
    global FIRST_RUN
    try:
        ou = adcontainer.ADContainer.from_dn(f'OU={ou_name},{p_dn}')
        logging.info(f'OU "{ou_name}" already exists')
        if ou_name == target and FIRST_RUN:
            FIRST_RUN = False
        return ou
    except Exception as e:
        parent_container = adcontainer.ADContainer.from_dn(p_dn)
        ou = parent_container.create_container(ou_name)
        logging.info(f'OU "{ou_name}" created successfully')
        return ou


def get_input_values():
    def close_handler():
        confirm = Messagebox.show_question(
            "Quit",
            "Do you want to quit?",
            parent=root
        )
        if confirm == "Yes":
            root.destroy()
            exit()

    # Default theme, will be changeable in form
    root = ttk.Window(themename='darkly')

    # Center the window on screen
    root.geometry('+500+150')
    input_form = InputForm(root)
    root.protocol("WM_DELETE_WINDOW", close_handler)

    root.mainloop()

    # Return the input values after the main loop has finished
    return {
        'ldap_server': input_form.ldap_server.get(),
        'username': input_form.username.get(),
        'password': input_form.password.get(),
        'source_file': input_form.source_file.get(),
        'destination_ou': input_form.destination_ou.get(),
        'domain': input_form.domain.get(),
        'upn_suffix': input_form.upn_suffix.get(),
        'result_file': input_form.result_file.get(),
        'logfile': input_form.logfile.get(),
        'protocol': input_form.protocol.get(),
        'theme': input_form.theme_var.get() if hasattr(input_form, 'theme_var') else 'darkly',
    }


# Main program start

if __name__ == '__main__':

    FIRST_RUN = True
    GROUP_YEAR = None

    input_data = get_input_values()

    logging.basicConfig(level=logging.INFO,
                        filename=f'{input_data['logfile']}.txt', filemode='w')

    logging.info(f'Import started at {datetime.now()}')
    logging.info(f'Data import from {input_data['source_file']} started')

    pyad_setdefaults(
        ldap_server=input_data['ldap_server'],
        username=input_data['username'],
        password=input_data['password'],
        ssl=input_data['protocol'] == 'LDAPS',
    )
    logging.info(
        f'LDAP server set to {input_data['ldap_server']} using {input_data['protocol']}')

    workbook = load_workbook(filename=f'{input_data['source_file']}')
    sheet = workbook.active
    logging.info(f'Workbook {input_data['source_file']} loaded')

    # Get data into variable

    user_data = []
    for row in sheet.iter_rows(max_col=8, min_row=2):
        cname, surname, passwd, full_name, phone, eng_surname, group_year, email = row
        if surname.value is not None:
            GROUP_YEAR = str(group_year.value)
            eng_surname = translit(
                str(surname.value), language_code='ru', reversed=True)
            cname = str(eng_surname) + str(group_year.value)
            email.value = ''.join([cname, '@', input_data['domain']])
            user_data.append({
                'cname': cname,
                'surname': surname.value,
                'passwd': passwd.value,
                'full_name': full_name.value,
                'phone': phone.value,
                'eng_surname': eng_surname,
                'group_year': group_year.value,
                'email': email.value})
        else:
            break

    logging.info(f'User data collected from excel workbook.')

    # Create OUs
    destination_ou_list = [
        value for value in input_data['destination_ou'].split('/') if value]
    parent_dn = f"DC={',DC='.join(input_data['domain'].split('.'))}"
    if destination_ou_list:
        target = destination_ou_list[-1]
        target_ou = {}
        for i, dn in enumerate(destination_ou_list):
            target_ou[f'{i}'] = dn
        if target_ou:
            for i in range(len(target_ou)):
                parent = get_or_create_ou(target_ou[f'{i}'], parent_dn)
                parent_dn = parent.dn
            user_creation_ou = parent
    else:
        user_creation_ou = adcontainer.ADContainer.from_dn(parent_dn)
        logging.info('Target OU for user creation is root.')

    # Insert data into AD and register success

    for i in range(0, len(user_data)):
        try:
            create_ad_user(ou=user_creation_ou,
                           upn_suffix=input_data['upn_suffix'], **user_data[i])
            sheet.cell(row=i + 2, column=9, value='Y')
            sheet.cell(row=i + 2, column=6, value=user_data[i]['eng_surname'])
        except Exception as e:
            sheet.cell(row=i + 2, column=9, value='N')
            sheet.cell(row=i + 2, column=10, value=f'Error {e}')
            logging.error(
                f'User {user_data[i]["cname"]} not created due to {e}.')

    workbook.save(filename=f'{input_data['result_file']}.xlsx')
